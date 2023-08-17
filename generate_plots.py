# module imports
import numpy as np
import matplotlib.pyplot as plt
import seaborn as sns
import pandas as pd
import argparse
import pprint

from openpyxl import load_workbook

# constants used in the script
file_path = './Questionnaire/Data/QuestionnaireData.xlsx'
NUM_QUESTIONS_TOTAL = 43
NUM_QUESTIONS_PRESENCE = 19
NUM_QUESTIONS_TLX = 6
NUM_QUESTIONS_EMBODIMENT = 18
NUM_QUESTIONS_EMBODIMENT_OWNERSHIP = 3
NUM_QUESTIONS_EMBODIMENT_AGENCY = 4
NUM_QUESTIONS_EMBODIMENT_TACTILE = 4
NUM_QUESTIONS_EMBODIMENT_LOCATION = 3
NUM_QUESTIONS_EMBODIMENT_APPEARANCE = 4

BASELINE = 0
TRAIN = 1
TEST = 2

def compute_embodiment_scores(with_haptics, without_haptics, experiment_block):
    # Scoring key: 
    # Ownership = Q1 - Q2 - Q3
    # Agency = Q4 + Q5 + Q6 - Q7
    # Tactile = Q8 - Q9 + Q10 + Q11
    # Location = Q12 (- Q13) + Q14
    # Appearance = Q15 + Q16 + Q17 + Q18

    # compute PCA scores for individual criteria
    ownership_score = [{'with_haptics': None, 'without_haptics': None}, {'with_haptics': None, 'without_haptics': None}, {'with_haptics': None, 'without_haptics': None}]
    agency_score = [{'with_haptics': None, 'without_haptics': None}, {'with_haptics': None, 'without_haptics': None}, {'with_haptics': None, 'without_haptics': None}]
    tactile_score = [{'with_haptics': None, 'without_haptics': None}, {'with_haptics': None, 'without_haptics': None}, {'with_haptics': None, 'without_haptics': None}]
    location_score = [{'with_haptics': None, 'without_haptics': None}, {'with_haptics': None, 'without_haptics': None}, {'with_haptics': None, 'without_haptics': None}]
    appearance_score = [{'with_haptics': None, 'without_haptics': None}, {'with_haptics': None, 'without_haptics': None}, {'with_haptics': None, 'without_haptics': None}]

    for block_id in range(experiment_block):
        # with_haptics
        ownership_score[block_id]['with_haptics'] = np.array(with_haptics[block_id]['Q1']) - np.array(with_haptics[block_id]['Q2']) - np.array(with_haptics[block_id]['Q3'])
        agency_score[block_id]['with_haptics'] = np.array(with_haptics[block_id]['Q4']) + np.array(with_haptics[block_id]['Q5']) + np.array(with_haptics[block_id]['Q6']) - np.array(with_haptics[block_id]['Q7'])
        tactile_score[block_id]['with_haptics'] = np.array(with_haptics[block_id]['Q8']) - np.array(with_haptics[block_id]['Q9']) + np.array(with_haptics[block_id]['Q10']) + np.array(with_haptics[block_id]['Q11'])
        location_score[block_id]['with_haptics'] = np.array(with_haptics[block_id]['Q12']) - np.array(with_haptics[block_id]['Q13']) + np.array(with_haptics[block_id]['Q14'])
        appearance_score[block_id]['with_haptics'] = np.array(with_haptics[block_id]['Q15']) + np.array(with_haptics[block_id]['Q16']) + np.array(with_haptics[block_id]['Q17']) + np.array(with_haptics[block_id]['Q18'])
        # without_haptics
        ownership_score[block_id]['without_haptics'] = np.array(without_haptics[block_id]['Q1']) - np.array(without_haptics[block_id]['Q2']) - np.array(without_haptics[block_id]['Q3'])
        agency_score[block_id]['without_haptics'] = np.array(without_haptics[block_id]['Q4']) + np.array(without_haptics[block_id]['Q5']) + np.array(without_haptics[block_id]['Q6']) - np.array(without_haptics[block_id]['Q7'])
        tactile_score[block_id]['without_haptics'] = np.array(without_haptics[block_id]['Q8']) - np.array(without_haptics[block_id]['Q9']) + np.array(without_haptics[block_id]['Q10']) + np.array(without_haptics[block_id]['Q11'])
        location_score[block_id]['without_haptics'] = np.array(without_haptics[block_id]['Q12']) - np.array(without_haptics[block_id]['Q13']) + np.array(without_haptics[block_id]['Q14'])
        appearance_score[block_id]['without_haptics'] = np.array(without_haptics[block_id]['Q15']) + np.array(without_haptics[block_id]['Q16']) + np.array(without_haptics[block_id]['Q17']) +  np.array(without_haptics[block_id]['Q18'])
        
    return ownership_score, agency_score, tactile_score, location_score, appearance_score 


def main(participant_id, experiment_block):
    # read the excel file for the questionnaire data
    file = load_workbook(file_path)
    
    #print(f'The number of sheets in this file is {len(file.sheetnames)}')
    
    if participant_id > len(file.sheetnames):
        print('Invalid participant ID!')
        return -1
    # identify the sheets from which data needs to be collected  
    all_data = pd.read_excel(file_path, sheet_name=None)
    # if a particular participant is selected, only choose that sheet 
    # while maintaining dict structure
    if participant_id != -1:
        for sheet_name, df in all_data.items():
            if str(participant_id) in sheet_name:
                all_data = {sheet_name: all_data[sheet_name]}
    
    #print(f'Specified ID: {participant_id}\nData:\n{all_data}')

    # create dictionaries to store questions and their score lists for each questionnaire
    presence_responses = [{}, {}, {}]
    for j in range(experiment_block):
        for i in range(1, NUM_QUESTIONS_PRESENCE + 1):
            presence_responses[j][f'Q{i}'] = []
    
    tlx_responses = [{}, {}, {}]
    for j in range(experiment_block):
        for i in range(1, NUM_QUESTIONS_TLX + 1):
            tlx_responses[j][f'Q{i}'] = []
    
    embodiment_responses = [{}, {}, {}]
    for j in range(experiment_block):
        for i in range(1, NUM_QUESTIONS_EMBODIMENT + 1):
            embodiment_responses[j][f'Q{i}'] = []

    # print(presence_responses)
    # print(tlx_responses)
    # print(embodiment_responses)    


    # now, regardless of how many participants we are plotting for,
    # iterate over the all_data object to accumulate scores
    count = 0
    # 1: WithHaptics participants
    for sheet_name, df in all_data.items():
        # check if the sheet is for a WithHaptics participant
        if 'WithHaptics' not in sheet_name:
            continue
        count += 1
        # iterate over experiment blocks to accumulate data
        for block_id in range(experiment_block):
            for question_id in range(1, NUM_QUESTIONS_PRESENCE + 1):
                presence_responses[block_id][f'Q{question_id}'].append(df.iloc[1 + block_id, 1 + question_id])
            for question_id in range(1, NUM_QUESTIONS_TLX + 1):
                tlx_responses[block_id][f'Q{question_id}'].append(df.iloc[5 + block_id, 1 + question_id])
            for question_id in range(1, NUM_QUESTIONS_EMBODIMENT_OWNERSHIP + 1):
                embodiment_responses[block_id][f'Q{question_id}'].append(df.iloc[9 + block_id, 2 + question_id])
            for question_id in range(1, NUM_QUESTIONS_EMBODIMENT_AGENCY + 1):
                embodiment_responses[block_id][f'Q{question_id + NUM_QUESTIONS_EMBODIMENT_OWNERSHIP}'].append(df.iloc[11 + 10 * block_id, 2 + question_id])
            for question_id in range(1, NUM_QUESTIONS_EMBODIMENT_TACTILE + 1):
                embodiment_responses[block_id][f'Q{question_id + NUM_QUESTIONS_EMBODIMENT_OWNERSHIP + NUM_QUESTIONS_EMBODIMENT_AGENCY}'].append(df.iloc[13 + 10 * block_id, 2 + question_id])
            for question_id in range(1, NUM_QUESTIONS_EMBODIMENT_LOCATION + 1):
                embodiment_responses[block_id][f'Q{question_id + NUM_QUESTIONS_EMBODIMENT_OWNERSHIP + NUM_QUESTIONS_EMBODIMENT_AGENCY + NUM_QUESTIONS_EMBODIMENT_TACTILE}'].append(df.iloc[15 + 10 * block_id, 2 + question_id])
            for question_id in range(1, NUM_QUESTIONS_EMBODIMENT_APPEARANCE + 1):
                embodiment_responses[block_id][f'Q{question_id + NUM_QUESTIONS_EMBODIMENT_OWNERSHIP + NUM_QUESTIONS_EMBODIMENT_AGENCY + NUM_QUESTIONS_EMBODIMENT_TACTILE + NUM_QUESTIONS_EMBODIMENT_LOCATION}'].append(df.iloc[17 + 10 * block_id, 2 + question_id])
    
    pp = pprint.PrettyPrinter(indent=3)

    print(f'Found {count} participant(s) under WithHaptics.')
    # print(f'Presence:')
    # pp.pprint(presence_responses)
    # print(f'TLX:')
    # pp.pprint(tlx_responses)
    # print(f'Embodiment:')
    # pp.pprint(embodiment_responses)

    # create dictionaries to store questions and their score lists for each questionnaire
    presence_responses_no = [{}, {}, {}]
    for j in range(experiment_block):
        for i in range(1, NUM_QUESTIONS_PRESENCE + 1):
            presence_responses_no[j][f'Q{i}'] = []
    
    tlx_responses_no = [{}, {}, {}]
    for j in range(experiment_block):
        for i in range(1, NUM_QUESTIONS_TLX + 1):
            tlx_responses_no[j][f'Q{i}'] = []
    
    embodiment_responses_no = [{}, {}, {}]
    for j in range(experiment_block):
        for i in range(1, NUM_QUESTIONS_EMBODIMENT + 1):
            embodiment_responses_no[j][f'Q{i}'] = []


    # now, regardless of how many participants we are plotting for,
    # iterate over the all_data object to accumulate scores
    count = 0
    # 1: WithHaptics participants
    for sheet_name, df in all_data.items():
        # check if the sheet is for a WithHaptics participant
        if 'WithoutHaptics' not in sheet_name:
            continue
        count += 1
        # iterate over experiment blocks to accumulate data
        for block_id in range(experiment_block):
            for question_id in range(1, NUM_QUESTIONS_PRESENCE + 1):
                presence_responses_no[block_id][f'Q{question_id}'].append(df.iloc[1 + block_id, 1 + question_id])
            for question_id in range(1, NUM_QUESTIONS_TLX + 1):
                tlx_responses_no[block_id][f'Q{question_id}'].append(df.iloc[5 + block_id, 1 + question_id])
            for question_id in range(1, NUM_QUESTIONS_EMBODIMENT_OWNERSHIP + 1):
                embodiment_responses_no[block_id][f'Q{question_id}'].append(df.iloc[9 + block_id, 2 + question_id])
            for question_id in range(1, NUM_QUESTIONS_EMBODIMENT_AGENCY + 1):
                embodiment_responses_no[block_id][f'Q{question_id + NUM_QUESTIONS_EMBODIMENT_OWNERSHIP}'].append(df.iloc[11 + 10 * block_id, 2 + question_id])
            for question_id in range(1, NUM_QUESTIONS_EMBODIMENT_TACTILE + 1):
                embodiment_responses_no[block_id][f'Q{question_id + NUM_QUESTIONS_EMBODIMENT_OWNERSHIP + NUM_QUESTIONS_EMBODIMENT_AGENCY}'].append(df.iloc[13 + 10 * block_id, 2 + question_id])
            for question_id in range(1, NUM_QUESTIONS_EMBODIMENT_LOCATION + 1):
                embodiment_responses_no[block_id][f'Q{question_id + NUM_QUESTIONS_EMBODIMENT_OWNERSHIP + NUM_QUESTIONS_EMBODIMENT_AGENCY + NUM_QUESTIONS_EMBODIMENT_TACTILE}'].append(df.iloc[15 + 10 * block_id, 2 + question_id])
            for question_id in range(1, NUM_QUESTIONS_EMBODIMENT_APPEARANCE + 1):
                embodiment_responses_no[block_id][f'Q{question_id + NUM_QUESTIONS_EMBODIMENT_OWNERSHIP + NUM_QUESTIONS_EMBODIMENT_AGENCY + NUM_QUESTIONS_EMBODIMENT_TACTILE + NUM_QUESTIONS_EMBODIMENT_LOCATION}'].append(df.iloc[17 + 10 * block_id, 2 + question_id])
    
    pp = pprint.PrettyPrinter(indent=3)

    print(f'Found {count} participant(s) under WithoutHaptics.')
    # print(f'Presence:')
    # pp.pprint(presence_responses_no)
    # print(f'TLX:')
    # pp.pprint(tlx_responses_no)
    # print(f'Embodiment:')
    # pp.pprint(embodiment_responses_no)

    print(f'Generating plots for Presence PCA...')
    sns.set_style(style='whitegrid')

    ownership_score, agency_score, tactile_score, location_score, appearance_score = compute_embodiment_scores(presence_responses, presence_responses_no, experiment_block)

    # move all data to pandas dataframes
    # df_presence = pd.DataFrame(presence_responses).dropna()
    # df_tlx = pd.DataFrame(tlx_responses).dropna()
    # df_embodiment = pd.DataFrame(embodiment_responses).dropna()
    # df_presence_no = pd.DataFrame(presence_responses_no[0]).dropna()
    # df_tlx_no = pd.DataFrame(tlx_responses_no).dropna()
    # df_embodiment_no = pd.DataFrame(embodiment_responses_no).dropna()

    custom_x_labels = ['Baseline', 'Train', 'Test']

    """
    1. Body Ownership Plot
    """

    # condition_labels = []
    # for i in range(len(ownership_score[BASELINE]['with_haptics'])):
    #     condition_labels.append(1)
    # phase_labels = []
    # for i in range(len(ownership_score[BASELINE]['with_haptics'])):
    #     phase_labels.append(0)
    # df_ownership_score_with_baseline = pd.DataFrame({'score': ownership_score[BASELINE]['with_haptics'], 'condition': condition_labels, 'phase': phase_labels})
    # condition_labels = []
    # for i in range(len(ownership_score[BASELINE]['without_haptics'])):
    #     condition_labels.append(0)
    # phase_labels = []
    # for i in range(len(ownership_score[BASELINE]['without_haptics'])):
    #     phase_labels.append(0)
    # df_ownership_score_without_baseline = pd.DataFrame({'score': ownership_score[BASELINE]['without_haptics'], 'condition': condition_labels, 'phase': phase_labels})
    # df_ownership_score_baseline = pd.concat([df_ownership_score_with_baseline, df_ownership_score_without_baseline])

    # condition_labels = []
    # for i in range(len(ownership_score[TRAIN]['with_haptics'])):
    #     condition_labels.append(1)
    # phase_labels = []
    # for i in range(len(ownership_score[TRAIN]['with_haptics'])):
    #     phase_labels.append(1)
    # df_ownership_score_with_train = pd.DataFrame({'score': ownership_score[TRAIN]['with_haptics'], 'condition': condition_labels, 'phase': phase_labels})
    # condition_labels = []
    # for i in range(len(ownership_score[TRAIN]['without_haptics'])):
    #     condition_labels.append(0)
    # phase_labels = []
    # for i in range(len(ownership_score[TRAIN]['without_haptics'])):
    #     phase_labels.append(1)
    # df_ownership_score_without_train = pd.DataFrame({'score': ownership_score[TRAIN]['without_haptics'], 'condition': condition_labels, 'phase': phase_labels})
    # df_ownership_score_train = pd.concat([df_ownership_score_with_train, df_ownership_score_without_train])

    # condition_labels = []
    # for i in range(len(ownership_score[TEST]['with_haptics'])):
    #     condition_labels.append(1)
    # phase_labels = []
    # for i in range(len(ownership_score[TEST]['with_haptics'])):
    #     phase_labels.append(2)
    # df_ownership_score_with_test = pd.DataFrame({'score': ownership_score[TEST]['with_haptics'], 'condition': condition_labels, 'phase': phase_labels})
    # condition_labels = []
    # for i in range(len(ownership_score[TEST]['without_haptics'])):
    #     condition_labels.append(0)
    # phase_labels = []
    # for i in range(len(ownership_score[TEST]['without_haptics'])):
    #     phase_labels.append(2)
    # df_ownership_score_without_test = pd.DataFrame({'score': ownership_score[TEST]['without_haptics'], 'condition': condition_labels, 'phase': phase_labels})
    # df_ownership_score_test = pd.concat([df_ownership_score_with_test, df_ownership_score_without_test])

    # df_ownership_score = pd.concat([df_ownership_score_baseline, df_ownership_score_train, df_ownership_score_test])

    # ax = sns.boxplot(data = df_ownership_score, x='phase', y='score', hue='condition')
    # ax.set(xlabel='Experiment Phase', ylabel='Body Ownership Score')
    # plotname = 'Embodiment-Ownership.png'

    """
    2. Agency and Motor Control Plot
    """

    # condition_labels = []
    # for i in range(len(agency_score[BASELINE]['with_haptics'])):
    #     condition_labels.append(1)
    # phase_labels = []
    # for i in range(len(agency_score[BASELINE]['with_haptics'])):
    #     phase_labels.append(0)
    # df_agency_score_with_baseline = pd.DataFrame({'score': agency_score[BASELINE]['with_haptics'], 'condition': condition_labels, 'phase': phase_labels})
    # condition_labels = []
    # for i in range(len(agency_score[BASELINE]['without_haptics'])):
    #     condition_labels.append(0)
    # phase_labels = []
    # for i in range(len(agency_score[BASELINE]['without_haptics'])):
    #     phase_labels.append(0)
    # df_agency_score_without_baseline = pd.DataFrame({'score': agency_score[BASELINE]['without_haptics'], 'condition': condition_labels, 'phase': phase_labels})
    # df_agency_score_baseline = pd.concat([df_agency_score_with_baseline, df_agency_score_without_baseline])

    # condition_labels = []
    # for i in range(len(agency_score[TRAIN]['with_haptics'])):
    #     condition_labels.append(1)
    # phase_labels = []
    # for i in range(len(agency_score[TRAIN]['with_haptics'])):
    #     phase_labels.append(1)
    # df_agency_score_with_train = pd.DataFrame({'score': agency_score[TRAIN]['with_haptics'], 'condition': condition_labels, 'phase': phase_labels})
    # condition_labels = []
    # for i in range(len(agency_score[TRAIN]['without_haptics'])):
    #     condition_labels.append(0)
    # phase_labels = []
    # for i in range(len(agency_score[TRAIN]['without_haptics'])):
    #     phase_labels.append(1)
    # df_agency_score_without_train = pd.DataFrame({'score': agency_score[TRAIN]['without_haptics'], 'condition': condition_labels, 'phase': phase_labels})
    # df_agency_score_train = pd.concat([df_agency_score_with_train, df_agency_score_without_train])

    # condition_labels = []
    # for i in range(len(agency_score[TEST]['with_haptics'])):
    #     condition_labels.append(1)
    # phase_labels = []
    # for i in range(len(agency_score[TEST]['with_haptics'])):
    #     phase_labels.append(2)
    # df_agency_score_with_test = pd.DataFrame({'score': agency_score[TEST]['with_haptics'], 'condition': condition_labels, 'phase': phase_labels})
    # condition_labels = []
    # for i in range(len(agency_score[TEST]['without_haptics'])):
    #     condition_labels.append(0)
    # phase_labels = []
    # for i in range(len(agency_score[TEST]['without_haptics'])):
    #     phase_labels.append(2)
    # df_agency_score_without_test = pd.DataFrame({'score': agency_score[TEST]['without_haptics'], 'condition': condition_labels, 'phase': phase_labels})
    # df_agency_score_test = pd.concat([df_agency_score_with_test, df_agency_score_without_test])

    # df_agency_score = pd.concat([df_agency_score_baseline, df_agency_score_train, df_agency_score_test])

    # ax = sns.boxplot(data = df_agency_score, x='phase', y='score', hue='condition')
    # ax.set(xlabel='Experiment Phase', ylabel='Agency Score')
    # plotname = 'Embodiment-Agency.png'

    """
    3. Tactile Perception Plot
    """

    # condition_labels = []
    # for i in range(len(tactile_score[BASELINE]['with_haptics'])):
    #     condition_labels.append(1)
    # phase_labels = []
    # for i in range(len(tactile_score[BASELINE]['with_haptics'])):
    #     phase_labels.append(0)
    # df_tactile_score_with_baseline = pd.DataFrame({'score': tactile_score[BASELINE]['with_haptics'], 'condition': condition_labels, 'phase': phase_labels})
    # condition_labels = []
    # for i in range(len(tactile_score[BASELINE]['without_haptics'])):
    #     condition_labels.append(0)
    # phase_labels = []
    # for i in range(len(tactile_score[BASELINE]['without_haptics'])):
    #     phase_labels.append(0)
    # df_tactile_score_without_baseline = pd.DataFrame({'score': tactile_score[BASELINE]['without_haptics'], 'condition': condition_labels, 'phase': phase_labels})
    # df_tactile_score_baseline = pd.concat([df_tactile_score_with_baseline, df_tactile_score_without_baseline])

    # condition_labels = []
    # for i in range(len(tactile_score[TRAIN]['with_haptics'])):
    #     condition_labels.append(1)
    # phase_labels = []
    # for i in range(len(tactile_score[TRAIN]['with_haptics'])):
    #     phase_labels.append(1)
    # df_tactile_score_with_train = pd.DataFrame({'score': tactile_score[TRAIN]['with_haptics'], 'condition': condition_labels, 'phase': phase_labels})
    # condition_labels = []
    # for i in range(len(tactile_score[TRAIN]['without_haptics'])):
    #     condition_labels.append(0)
    # phase_labels = []
    # for i in range(len(tactile_score[TRAIN]['without_haptics'])):
    #     phase_labels.append(1) 
    # df_tactile_score_without_train = pd.DataFrame({'score': tactile_score[TRAIN]['without_haptics'], 'condition': condition_labels, 'phase': phase_labels})
    # df_tactile_score_train = pd.concat([df_tactile_score_with_train, df_tactile_score_without_train])

    # condition_labels = []
    # for i in range(len(tactile_score[TEST]['with_haptics'])):
    #     condition_labels.append(1)
    # phase_labels = []
    # for i in range(len(tactile_score[TEST]['with_haptics'])):
    #     phase_labels.append(2)
    # df_tactile_score_with_test = pd.DataFrame({'score': tactile_score[TEST]['with_haptics'], 'condition': condition_labels, 'phase': phase_labels})
    # condition_labels = []
    # for i in range(len(tactile_score[TEST]['without_haptics'])):
    #     condition_labels.append(0)
    # phase_labels = []
    # for i in range(len(tactile_score[TEST]['without_haptics'])):
    #     phase_labels.append(2)
    # df_tactile_score_without_test = pd.DataFrame({'score': tactile_score[TEST]['without_haptics'], 'condition': condition_labels, 'phase': phase_labels})
    # df_tactile_score_test = pd.concat([df_tactile_score_with_test, df_tactile_score_without_test])

    # df_tactile_score = pd.concat([df_tactile_score_baseline, df_tactile_score_train, df_tactile_score_test])

    # ax = sns.boxplot(data=df_tactile_score, x='phase', y='score', hue='condition')
    # ax.set(xlabel='Experiment Phase', ylabel='Tactile Perception Score')
    # plotname = 'Embodiment-Tactile.png'

    """
    4. Location Plot
    """

    # condition_labels = []
    # for i in range(len(location_score[BASELINE]['with_haptics'])):
    #     condition_labels.append(1)
    # phase_labels = []
    # for i in range(len(location_score[BASELINE]['with_haptics'])):
    #     phase_labels.append(0)
    # df_location_score_with_baseline = pd.DataFrame({'score': location_score[BASELINE]['with_haptics'], 'condition': condition_labels, 'phase': phase_labels})
    # condition_labels = []
    # for i in range(len(location_score[BASELINE]['without_haptics'])):
    #     condition_labels.append(0)
    # phase_labels = []
    # for i in range(len(location_score[BASELINE]['without_haptics'])):
    #     phase_labels.append(0)
    # df_location_score_without_baseline = pd.DataFrame({'score': location_score[BASELINE]['without_haptics'], 'condition': condition_labels, 'phase': phase_labels})
    # df_location_score_baseline = pd.concat([df_location_score_with_baseline, df_location_score_without_baseline])

    # condition_labels = []
    # for i in range(len(location_score[TRAIN]['with_haptics'])):
    #     condition_labels.append(1)
    # phase_labels = []
    # for i in range(len(location_score[TRAIN]['with_haptics'])):
    #     phase_labels.append(1) 
    # df_location_score_with_train = pd.DataFrame({'score': location_score[TRAIN]['with_haptics'], 'condition': condition_labels, 'phase': phase_labels})
    # condition_labels = []
    # for i in range(len(location_score[TRAIN]['without_haptics'])):
    #     condition_labels.append(0)
    # phase_labels = []
    # for i in range(len(location_score[TRAIN]['without_haptics'])):
    #     phase_labels.append(1)
    # df_location_score_without_train = pd.DataFrame({'score': location_score[TRAIN]['without_haptics'], 'condition': condition_labels, 'phase': phase_labels})
    # df_location_score_train = pd.concat([df_location_score_with_train, df_location_score_without_train])

    # condition_labels = []
    # for i in range(len(location_score[TEST]['with_haptics'])):
    #     condition_labels.append(1)
    # phase_labels = []
    # for i in range(len(location_score[TEST]['with_haptics'])):
    #     phase_labels.append(2) 
    # df_location_score_with_test = pd.DataFrame({'score': location_score[TEST]['with_haptics'], 'condition': condition_labels, 'phase': phase_labels})
    # condition_labels = []
    # for i in range(len(location_score[TEST]['without_haptics'])):
    #     condition_labels.append(0)
    # phase_labels = []
    # for i in range(len(location_score[TEST]['without_haptics'])):
    #     phase_labels.append(2) 
    # df_location_score_without_test = pd.DataFrame({'score': location_score[TEST]['without_haptics'], 'condition': condition_labels, 'phase': phase_labels})
    # df_location_score_test = pd.concat([df_location_score_with_test, df_location_score_without_test])

    # df_location_score = pd.concat([df_location_score_baseline, df_location_score_train, df_location_score_test])

    # ax = sns.boxplot(data=df_location_score, x='phase', y='score', hue='condition')
    # ax.set(xlabel='Experiment Phase', ylabel='Location Score')
    # plotname = 'Embodiment-Location.png'

    """
    5. Appearance Plot
    """

    condition_labels = []
    for i in range(len(appearance_score[BASELINE]['with_haptics'])):
        condition_labels.append(1)
    phase_labels = []
    for i in range(len(appearance_score[BASELINE]['with_haptics'])):
        phase_labels.append(0)
    df_appearance_score_with_baseline = pd.DataFrame({'score': appearance_score[BASELINE]['with_haptics'], 'condition': condition_labels, 'phase': phase_labels})
    condition_labels = []
    for i in range(len(appearance_score[BASELINE]['without_haptics'])):
        condition_labels.append(0)
    phase_labels = []
    for i in range(len(appearance_score[BASELINE]['without_haptics'])):
        phase_labels.append(0)
    df_appearance_score_without_baseline = pd.DataFrame({'score': appearance_score[BASELINE]['without_haptics'], 'condition': condition_labels, 'phase': phase_labels})
    df_appearance_score_baseline = pd.concat([df_appearance_score_with_baseline, df_appearance_score_without_baseline])

    condition_labels = []
    for i in range(len(appearance_score[TRAIN]['with_haptics'])):
        condition_labels.append(1)
    phase_labels = []
    for i in range(len(appearance_score[TRAIN]['with_haptics'])):
        phase_labels.append(1) 
    df_appearance_score_with_train = pd.DataFrame({'score': appearance_score[TRAIN]['with_haptics'], 'condition': condition_labels, 'phase': phase_labels})
    condition_labels = []
    for i in range(len(appearance_score[TRAIN]['without_haptics'])):
        condition_labels.append(0)
    phase_labels = []
    for i in range(len(appearance_score[TRAIN]['without_haptics'])):
        phase_labels.append(1)
    df_appearance_score_without_train = pd.DataFrame({'score': appearance_score[TRAIN]['without_haptics'], 'condition': condition_labels, 'phase': phase_labels})
    df_appearance_score_train = pd.concat([df_appearance_score_with_train, df_appearance_score_without_train])

    condition_labels = []
    for i in range(len(appearance_score[TEST]['with_haptics'])):
        condition_labels.append(1)
    phase_labels = []
    for i in range(len(appearance_score[TEST]['with_haptics'])):
        phase_labels.append(2) 
    df_appearance_score_with_test = pd.DataFrame({'score': appearance_score[TEST]['with_haptics'], 'condition': condition_labels, 'phase': phase_labels})
    condition_labels = []
    for i in range(len(appearance_score[TEST]['without_haptics'])):
        condition_labels.append(0)
    phase_labels = []
    for i in range(len(appearance_score[TEST]['without_haptics'])):
        phase_labels.append(2) 
    df_appearance_score_without_test = pd.DataFrame({'score': appearance_score[TEST]['without_haptics'], 'condition': condition_labels, 'phase': phase_labels})
    df_appearance_score_test = pd.concat([df_appearance_score_with_test, df_appearance_score_without_test])

    df_appearance_score = pd.concat([df_appearance_score_baseline, df_appearance_score_train, df_appearance_score_test])

    ax = sns.boxplot(data=df_appearance_score, x='phase', y='score', hue='condition')
    ax.set(xlabel='Experiment Phase', ylabel='Appearance Score')
    plotname = 'Embodiment-Appearance.png'   
    
    
    ax.set_xticklabels(custom_x_labels)
    plt.savefig(plotname)

    # df_appearance_score = pd.concat([df_appearance_score_with, df_appearance_score_without])

    # custom_x_labels = ['Ownership-WithHaptics', 'Ownership-WithoutHaptics', 'Agency-WithHaptics', 'Agency-WithoutHaptics', 'Tactile-WithHaptics', 'Tactile-WithoutHaptics', 'Location-WithHaptics', 'Location-WithoutHaptics', 'Appearance-WithHaptics', 'Appearance-WithoutHaptics']

    #ax = sns.boxplot(data=[df_ownership_score_with, df_appearance_score_without, df_agency_score_with, df_agency_score_without, df_tactile_score_with, df_tactile_score_without, df_location_score_with, df_location_score_without, df_appearance_score_with, df_appearance_score_without])

    #ax.set_xticklabels(custom_x_labels)


    # plt.show()

    


if __name__ == "__main__":
    # Create an argument parser
    parser = argparse.ArgumentParser(description="Data analysis script for the Mirror Therapy questionnaires. The questionnaires \
                                     considered for the purposes of the experiment include Presence, NASA TLX and Avatar Embodiment.")

    # Add arguments to the parser
    parser.add_argument("participant_id", help="Integer indicating the ID of the participant for which plots are to be generated.\
                                            If all participants are considered, enter '-1'.")
    parser.add_argument("experiment_block", help="Integer indicating the experiment phase up till which plots are to be generated (1, 2, 3).")

    # Parse the command-line arguments
    args = parser.parse_args()

    # Call the main function with the provided arguments
    main(int(args.participant_id), int(args.experiment_block))